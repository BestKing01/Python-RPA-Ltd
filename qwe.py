import os
from openpyxl import *

def save_to_excel(directory, workbook, sheet):
    row = 1
    for root, dirs, files in os.walk(directory):
        for file in files:
            sheet.cell(row, 1, row)
            sheet.cell(row, 2, root)
            sheet.cell(row, 3, file)
            try:
                file_extension = file.split(".")[-1]
            except IndexError:
                file_extension = "None"
            sheet.cell(row, 4, file_extension)
            row += 1

if __name__ == "__main__":
    try:    
        directory = input("Directory: ")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "File List"
        sheet.cell(1, 1, "Line number")
        sheet.cell(1, 2, "Folder where the file is located")
        sheet.cell(1, 3, "File name")
        sheet.cell(1, 4, "File extension")
        save_to_excel(directory, workbook, sheet)
        workbook.save(filename="result.xlsx")
        print("File saved successfully.")
    except Exception as e:
        print("An error occurred:", e)

